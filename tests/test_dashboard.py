# MTR DUAT - Dashboard Analysis Tests
"""Unit tests for analysis/dashboard.py"""

import pytest
import pandas as pd
from datetime import datetime

from analysis.dashboard import (
    aggregate_records,
    calculate_summary,
    get_weekly_trend,
    get_monthly_trend,
    get_project_distribution,
    get_keyword_distribution,
    get_nth_pivot_by_week,
    export_dashboard_excel,
    DashboardAnalyzer,
)


def _make_records(n: int = 6):
    """Helper: build a list of realistic record dicts.

    Note: FullDate contains "Day dd/mm" (no year).  aggregate_records
    appends the Year column to build the full date string for parsing.
    """
    base = [
        {"FullDate": "Mon 01/01", "Project": "C2264", "Qty Delivered": 5, "Week": "WK01", "Year": "2024", "Line": "EAL"},
        {"FullDate": "Tue 02/01", "Project": "C2264", "Qty Delivered": 3, "Week": "WK01", "Year": "2024", "Line": "EAL"},
        {"FullDate": "Wed 03/01", "Project": "CBM", "Qty Delivered": 2, "Week": "WK01", "Year": "2024", "Line": "EAL"},
        {"FullDate": "Thu 08/01", "Project": "C2264", "Qty Delivered": 4, "Week": "WK02", "Year": "2024", "Line": "EAL"},
        {"FullDate": "Fri 09/01", "Project": "PA work", "Qty Delivered": 1, "Week": "WK02", "Year": "2024", "Line": "EAL"},
        {"FullDate": "Sat 10/01", "Project": "Provide", "Qty Delivered": 6, "Week": "WK02", "Year": "2024", "Line": "EAL"},
    ]
    return base[:n]


# ── aggregate_records ────────────────────────────────────────────────────────


@pytest.mark.unit
class TestAggregateRecords:

    def test_basic_aggregation(self):
        records = _make_records()
        df = aggregate_records(records)

        assert not df.empty
        assert len(df) == 6
        assert "Day" in df.columns
        assert "Date" in df.columns
        assert "DateObj" in df.columns
        assert "Month" in df.columns

    def test_day_extraction(self):
        records = _make_records(1)
        df = aggregate_records(records)
        assert df.iloc[0]["Day"] == "Mon"

    def test_week_numeric_extraction(self):
        records = _make_records(1)
        df = aggregate_records(records)
        assert df.iloc[0]["Week"] == "01"

    def test_empty_records_returns_empty_df(self):
        df = aggregate_records([])
        assert isinstance(df, pd.DataFrame)
        assert df.empty


# ── calculate_summary ────────────────────────────────────────────────────────


@pytest.mark.unit
class TestCalculateSummary:

    def test_produces_correct_columns(self):
        df = aggregate_records(_make_records())
        summary = calculate_summary(df, current_week=2, current_month=1)

        expected_cols = {"Rank", "Project", "Qty Delivered", "Total NTH",
                         "Qty per NTH", "Avg Qty per Week", "Avg Qty per Month",
                         "Avg NTH per Week", "Avg NTH per Month"}
        assert expected_cols.issubset(set(summary.columns))

    def test_rank_starts_at_one(self):
        df = aggregate_records(_make_records())
        summary = calculate_summary(df, current_week=2, current_month=1)
        assert summary.iloc[0]["Rank"] == 1

    def test_total_nth_counts_rows(self):
        df = aggregate_records(_make_records())
        summary = calculate_summary(df, current_week=2, current_month=1)
        c2264_row = summary[summary["Project"] == "C2264"].iloc[0]
        assert c2264_row["Total NTH"] == 3  # 3 records for C2264

    def test_empty_df_returns_empty(self):
        result = calculate_summary(pd.DataFrame())
        assert isinstance(result, pd.DataFrame)
        assert result.empty


# ── get_weekly_trend ─────────────────────────────────────────────────────────


@pytest.mark.unit
class TestGetWeeklyTrend:

    def test_returns_labels_and_categories(self):
        df = aggregate_records(_make_records())
        labels, data = get_weekly_trend(df)

        assert isinstance(labels, list)
        assert len(labels) > 0
        assert "Projects" in data
        assert "Jobs" in data

    def test_empty_df_returns_empty(self):
        labels, data = get_weekly_trend(pd.DataFrame())
        assert labels == []
        assert data == {}

    def test_none_df_returns_empty(self):
        labels, data = get_weekly_trend(None)
        assert labels == []
        assert data == {}


# ── get_monthly_trend ────────────────────────────────────────────────────────


@pytest.mark.unit
class TestGetMonthlyTrend:

    def test_returns_month_labels_and_counts(self):
        df = aggregate_records(_make_records())
        labels, counts = get_monthly_trend(df)

        assert isinstance(labels, list)
        assert isinstance(counts, list)
        assert len(labels) == len(counts)
        assert all(isinstance(c, int) for c in counts)

    def test_empty_df_returns_empty(self):
        labels, counts = get_monthly_trend(pd.DataFrame())
        assert labels == []
        assert counts == []


# ── get_project_distribution ─────────────────────────────────────────────────


@pytest.mark.unit
class TestGetProjectDistribution:

    def test_returns_project_count_dict(self):
        df = aggregate_records(_make_records())
        dist = get_project_distribution(df)

        assert isinstance(dist, dict)
        assert "C2264" in dist
        assert dist["C2264"] == 3

    def test_empty_df_returns_empty(self):
        assert get_project_distribution(pd.DataFrame()) == {}

    def test_provide_renamed(self):
        df = aggregate_records(_make_records())
        dist = get_project_distribution(df)
        assert "Provide manpower for switching" in dist


# ── get_keyword_distribution ─────────────────────────────────────────────────


@pytest.mark.unit
class TestGetKeywordDistribution:

    def test_returns_keyword_count_dict(self):
        df = aggregate_records(_make_records())
        dist = get_keyword_distribution(df)

        assert isinstance(dist, dict)
        assert "CBM" in dist
        assert dist["CBM"] == 1

    def test_provide_display_name(self):
        df = aggregate_records(_make_records())
        dist = get_keyword_distribution(df)
        assert "Provide manpower for switching" in dist

    def test_empty_df_returns_empty(self):
        assert get_keyword_distribution(pd.DataFrame()) == {}


# ── get_nth_pivot_by_week ────────────────────────────────────────────────────


@pytest.mark.unit
class TestGetNthPivotByWeek:

    def test_creates_pivot_table(self):
        df = aggregate_records(_make_records())
        pivot = get_nth_pivot_by_week(df)

        assert isinstance(pivot, pd.DataFrame)
        assert not pivot.empty
        assert "C2264" in pivot.columns

    def test_empty_df_returns_empty(self):
        result = get_nth_pivot_by_week(pd.DataFrame())
        assert result.empty


# ── DashboardAnalyzer ────────────────────────────────────────────────────────


@pytest.mark.unit
class TestDashboardAnalyzer:

    def test_load_from_records(self):
        analyzer = DashboardAnalyzer()
        result = analyzer.load_from_records(_make_records(), max_week=2)

        assert result is True
        assert analyzer.df is not None
        assert analyzer.summary is not None

    def test_get_stats_returns_correct_keys(self):
        analyzer = DashboardAnalyzer()
        analyzer.load_from_records(_make_records(), max_week=2)
        stats = analyzer.get_stats()

        expected_keys = {"total_records", "total_nth", "total_qty",
                         "unique_projects", "last_updated"}
        assert expected_keys == set(stats.keys())
        assert stats["total_records"] == 6
        assert stats["unique_projects"] == 4

    def test_empty_records_returns_false(self):
        analyzer = DashboardAnalyzer()
        result = analyzer.load_from_records([])
        assert result is False

    def test_get_stats_before_load_returns_empty(self):
        analyzer = DashboardAnalyzer()
        assert analyzer.get_stats() == {}


# ── export_dashboard_excel ──────────────────────────────────────────────────


@pytest.mark.unit
class TestExportDashboardExcel:

    def test_creates_excel_with_three_sheets(self, tmp_path):
        df = aggregate_records(_make_records())
        summary = calculate_summary(df, current_week=2, current_month=1)
        nth_pivot = get_nth_pivot_by_week(df)

        raw_df = df[["Year", "Day", "Date", "Week", "Project", "Qty Delivered"]].copy()
        output = tmp_path / "dashboard.xlsx"

        result = export_dashboard_excel(raw_df, summary, nth_pivot, output)

        assert result is True
        assert output.exists()

        # Verify sheets
        from openpyxl import load_workbook
        wb = load_workbook(output)
        sheet_names = wb.sheetnames
        assert "Raw Data" in sheet_names
        assert "Summary by Project" in sheet_names
        assert "NTH Trend by Week" in sheet_names
        wb.close()

    def test_with_empty_nth_pivot(self, tmp_path):
        df = aggregate_records(_make_records())
        summary = calculate_summary(df, current_week=2, current_month=1)
        raw_df = df[["Year", "Day", "Date", "Week", "Project", "Qty Delivered"]].copy()
        output = tmp_path / "dashboard_no_pivot.xlsx"

        result = export_dashboard_excel(raw_df, summary, pd.DataFrame(), output)

        assert result is True
        assert output.exists()


# ── DashboardAnalyzer.load_from_excel ───────────────────────────────────────


@pytest.mark.unit
class TestDashboardAnalyzerLoadFromExcel:

    def test_load_from_excel_with_valid_file(self, tmp_path):
        # First create a valid Excel file via export
        df = aggregate_records(_make_records())
        summary = calculate_summary(df, current_week=2, current_month=1)
        nth_pivot = get_nth_pivot_by_week(df)
        raw_df = df[["Year", "Day", "Date", "Week", "Project", "Qty Delivered"]].copy()

        filepath = tmp_path / "dashboard_load.xlsx"
        export_dashboard_excel(raw_df, summary, nth_pivot, filepath)

        # Now load it back
        analyzer = DashboardAnalyzer()
        result = analyzer.load_from_excel(filepath)

        assert result is True
        assert analyzer.df is not None
        assert analyzer.summary is not None
        assert analyzer.last_updated is not None

    def test_load_from_excel_invalid_file(self, tmp_path):
        filepath = tmp_path / "nonexistent.xlsx"
        analyzer = DashboardAnalyzer()
        result = analyzer.load_from_excel(filepath)
        assert result is False


# ── DashboardAnalyzer.export ────────────────────────────────────────────────


@pytest.mark.unit
class TestDashboardAnalyzerExport:

    def test_export_creates_excel_file(self, tmp_path):
        analyzer = DashboardAnalyzer()
        analyzer.load_from_records(_make_records(), max_week=2)

        output = tmp_path / "dashboard_export.xlsx"
        result = analyzer.export(output)

        assert result is True
        assert output.exists()

        # Verify content
        from openpyxl import load_workbook
        wb = load_workbook(output)
        assert "Raw Data" in wb.sheetnames
        assert "Summary by Project" in wb.sheetnames
        wb.close()

    def test_export_returns_false_without_data(self, tmp_path):
        analyzer = DashboardAnalyzer()
        output = tmp_path / "empty_export.xlsx"
        result = analyzer.export(output)
        assert result is False


# ── get_weekly_trend with Date column fallback ──────────────────────────────


@pytest.mark.unit
class TestGetWeeklyTrendDateFallback:

    def test_with_date_column_fallback(self):
        """When Year/Week columns are absent, should fall back to Date column."""
        df = pd.DataFrame([
            {"Project": "C2264", "Date": "01/01/2024", "Qty Delivered": 5},
            {"Project": "C2264", "Date": "08/01/2024", "Qty Delivered": 3},
            {"Project": "CBM", "Date": "15/01/2024", "Qty Delivered": 2},
        ])

        labels, data = get_weekly_trend(df)

        assert isinstance(labels, list)
        assert len(labels) > 0
        assert "Projects" in data
        assert "Jobs" in data


# ── get_monthly_trend with Date column fallback ─────────────────────────────


@pytest.mark.unit
class TestGetMonthlyTrendDateFallback:

    def test_with_date_column_fallback(self):
        """When Year/Week columns are absent, should fall back to Date column."""
        df = pd.DataFrame([
            {"Project": "C2264", "Date": "01/01/2024", "Qty Delivered": 5},
            {"Project": "C2264", "Date": "15/02/2024", "Qty Delivered": 3},
            {"Project": "CBM", "Date": "20/03/2024", "Qty Delivered": 2},
        ])

        labels, counts = get_monthly_trend(df)

        assert isinstance(labels, list)
        assert len(labels) > 0
        assert isinstance(counts, list)
        assert len(labels) == len(counts)
