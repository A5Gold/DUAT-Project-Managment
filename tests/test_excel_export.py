# MTR DUAT - Excel Export Module Tests
"""TDD tests for utils/excel_export.py - written BEFORE implementation."""

import pytest
import pandas as pd
import openpyxl
from pathlib import Path

from utils.excel_export import (
    export_dataframe_to_excel,
    create_dashboard_excel,
    export_lag_analysis_report,
)


# ---------------------------------------------------------------------------
# Fixtures
# ---------------------------------------------------------------------------

@pytest.fixture
def sample_df() -> pd.DataFrame:
    """A small DataFrame for basic export tests."""
    return pd.DataFrame({
        "Project": ["Alpha", "Beta", "Gamma"],
        "Qty Delivered": [10, 20, 30],
        "Week": [1, 2, 3],
    })


@pytest.fixture
def empty_df() -> pd.DataFrame:
    """An empty DataFrame (columns but no rows)."""
    return pd.DataFrame(columns=["A", "B", "C"])


@pytest.fixture
def raw_df() -> pd.DataFrame:
    """Raw data DataFrame matching dashboard export expectations."""
    return pd.DataFrame({
        "Year": [2025, 2025],
        "Day": ["Mon", "Tue"],
        "Date": ["01/01", "02/01"],
        "Week": [1, 1],
        "Project": ["Alpha", "Beta"],
        "Qty Delivered": [5, 10],
    })


@pytest.fixture
def summary_df() -> pd.DataFrame:
    """Summary DataFrame matching dashboard export expectations."""
    return pd.DataFrame({
        "Rank": [1, 2],
        "Project": ["Beta", "Alpha"],
        "Qty Delivered": [10, 5],
        "Total NTH": [1, 1],
    })


@pytest.fixture
def nth_trend_df() -> pd.DataFrame:
    """NTH trend pivot table (index = YearWeek, columns = projects)."""
    return pd.DataFrame(
        {"Alpha": [1, 0], "Beta": [0, 1]},
        index=pd.Index(["2025-W01", "2025-W02"], name="YearWeek"),
    )


@pytest.fixture
def lag_results_df() -> pd.DataFrame:
    """Lag analysis results DataFrame including Status Color column."""
    return pd.DataFrame({
        "Project No": ["C1234", "C5678"],
        "Project": ["Alpha", "Beta"],
        "NTH Lag/Lead": [-5, 3],
        "Status": ["lag_behind", "lag_on_track"],
        "Status Color": ["orange", "green"],
    })


# ===========================================================================
# Tests: export_dataframe_to_excel
# ===========================================================================

class TestExportDataframeToExcel:
    """Tests for the basic single-sheet export helper."""

    @pytest.mark.unit
    def test_basic_export_returns_true(self, tmp_path: Path, sample_df: pd.DataFrame):
        """Exporting a normal DataFrame returns True."""
        path = tmp_path / "out.xlsx"
        result = export_dataframe_to_excel(sample_df, path)
        assert result is True

    @pytest.mark.unit
    def test_basic_export_creates_file(self, tmp_path: Path, sample_df: pd.DataFrame):
        """The output file must exist after a successful export."""
        path = tmp_path / "out.xlsx"
        export_dataframe_to_excel(sample_df, path)
        assert path.exists()

    @pytest.mark.unit
    def test_basic_export_valid_xlsx(self, tmp_path: Path, sample_df: pd.DataFrame):
        """The output file must be a valid xlsx readable by openpyxl."""
        path = tmp_path / "out.xlsx"
        export_dataframe_to_excel(sample_df, path)
        wb = openpyxl.load_workbook(path)
        ws = wb.active
        # Header row + 3 data rows
        assert ws.max_row == 4
        assert ws.max_column == 3
        wb.close()

    @pytest.mark.unit
    def test_empty_df_returns_true(self, tmp_path: Path, empty_df: pd.DataFrame):
        """Exporting an empty DataFrame still returns True (empty sheet)."""
        path = tmp_path / "empty.xlsx"
        result = export_dataframe_to_excel(empty_df, path)
        assert result is True

    @pytest.mark.unit
    def test_empty_df_creates_file(self, tmp_path: Path, empty_df: pd.DataFrame):
        """An empty DataFrame still produces a valid file."""
        path = tmp_path / "empty.xlsx"
        export_dataframe_to_excel(empty_df, path)
        assert path.exists()

    @pytest.mark.unit
    def test_invalid_path_returns_false(self, sample_df: pd.DataFrame):
        """An unreachable path must return False, not raise."""
        bad_path = Path("Z:/nonexistent/dir/sub/file.xlsx")
        result = export_dataframe_to_excel(sample_df, bad_path)
        assert result is False


# ===========================================================================
# Tests: create_dashboard_excel
# ===========================================================================

class TestCreateDashboardExcel:
    """Tests for the multi-sheet dashboard export."""

    @pytest.mark.unit
    def test_three_sheets_with_nth_trend(
        self,
        tmp_path: Path,
        raw_df: pd.DataFrame,
        summary_df: pd.DataFrame,
        nth_trend_df: pd.DataFrame,
    ):
        """When nth_trend is provided, the workbook must have 3 sheets."""
        path = tmp_path / "dashboard.xlsx"
        result = create_dashboard_excel(raw_df, summary_df, path, nth_trend_df)
        assert result is True

        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Raw Data", "Summary by Project", "NTH Trend by Week"]
        wb.close()

    @pytest.mark.unit
    def test_two_sheets_without_nth_trend(
        self,
        tmp_path: Path,
        raw_df: pd.DataFrame,
        summary_df: pd.DataFrame,
    ):
        """When nth_trend is None, only 2 sheets should be created."""
        path = tmp_path / "dashboard_no_nth.xlsx"
        result = create_dashboard_excel(raw_df, summary_df, path, nth_trend=None)
        assert result is True

        wb = openpyxl.load_workbook(path)
        assert wb.sheetnames == ["Raw Data", "Summary by Project"]
        wb.close()

    @pytest.mark.unit
    def test_raw_data_content(
        self,
        tmp_path: Path,
        raw_df: pd.DataFrame,
        summary_df: pd.DataFrame,
    ):
        """Raw Data sheet must contain the correct header and row count."""
        path = tmp_path / "dashboard_content.xlsx"
        create_dashboard_excel(raw_df, summary_df, path)

        wb = openpyxl.load_workbook(path)
        ws = wb["Raw Data"]
        headers = [cell.value for cell in ws[1]]
        assert "Project" in headers
        assert "Qty Delivered" in headers
        # header + 2 data rows
        assert ws.max_row == 3
        wb.close()

    @pytest.mark.unit
    def test_invalid_path_returns_false(
        self,
        raw_df: pd.DataFrame,
        summary_df: pd.DataFrame,
    ):
        """An unreachable path must return False."""
        bad_path = Path("Z:/nonexistent/dir/dashboard.xlsx")
        result = create_dashboard_excel(raw_df, summary_df, bad_path)
        assert result is False


# ===========================================================================
# Tests: export_lag_analysis_report
# ===========================================================================

class TestExportLagAnalysisReport:
    """Tests for the lag analysis Excel export."""

    @pytest.mark.unit
    def test_creates_lag_analysis_sheet(
        self, tmp_path: Path, lag_results_df: pd.DataFrame
    ):
        """Output must contain a sheet named 'Lag Analysis'."""
        path = tmp_path / "lag.xlsx"
        result = export_lag_analysis_report(lag_results_df, path)
        assert result is True

        wb = openpyxl.load_workbook(path)
        assert "Lag Analysis" in wb.sheetnames
        wb.close()

    @pytest.mark.unit
    def test_drops_status_color_column(
        self, tmp_path: Path, lag_results_df: pd.DataFrame
    ):
        """The 'Status Color' column must NOT appear in the exported sheet."""
        path = tmp_path / "lag_no_color.xlsx"
        export_lag_analysis_report(lag_results_df, path)

        wb = openpyxl.load_workbook(path)
        ws = wb["Lag Analysis"]
        headers = [cell.value for cell in ws[1]]
        assert "Status Color" not in headers
        # Other columns must still be present
        assert "Project No" in headers
        assert "Status" in headers
        wb.close()

    @pytest.mark.unit
    def test_auto_adjusts_column_widths(
        self, tmp_path: Path, lag_results_df: pd.DataFrame
    ):
        """Column widths should be adjusted (wider than the default 8)."""
        path = tmp_path / "lag_widths.xlsx"
        export_lag_analysis_report(lag_results_df, path)

        wb = openpyxl.load_workbook(path)
        ws = wb["Lag Analysis"]
        # At least one column should be wider than the openpyxl default (~8)
        widths = [
            ws.column_dimensions[col].width
            for col in ws.column_dimensions
            if ws.column_dimensions[col].width is not None
        ]
        assert len(widths) > 0
        assert any(w > 8 for w in widths)
        wb.close()

    @pytest.mark.unit
    def test_invalid_path_returns_false(self, lag_results_df: pd.DataFrame):
        """An unreachable path must return False."""
        bad_path = Path("Z:/nonexistent/dir/lag.xlsx")
        result = export_lag_analysis_report(lag_results_df, bad_path)
        assert result is False
