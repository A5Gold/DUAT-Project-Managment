# MTR DUAT - Lag Analysis Tests
"""Unit tests for analysis/lag_analysis.py"""

import pytest
import pandas as pd
from datetime import datetime, timedelta
from pathlib import Path

from analysis.lag_analysis import (
    get_status,
    load_project_master,
    calculate_nth_lag_lead,
    export_lag_report,
    LagAnalyzer,
)


# ── get_status ───────────────────────────────────────────────────────────────


@pytest.mark.unit
class TestGetStatus:

    def test_urgent_lag(self):
        status, color = get_status(-15)
        assert status == "lag_urgent"
        assert color == "red"

    def test_behind_lag(self):
        status, color = get_status(-7)
        assert status == "lag_behind"
        assert color == "orange"

    def test_slight_lag(self):
        status, color = get_status(-2)
        assert status == "lag_slight"
        assert color == "yellow"

    def test_on_track(self):
        status, color = get_status(3)
        assert status == "lag_on_track"
        assert color == "green"

    def test_ahead(self):
        status, color = get_status(10)
        assert status == "lag_ahead"
        assert color == "blue"

    def test_boundary_minus_10(self):
        status, color = get_status(-10)
        assert status == "lag_urgent"
        assert color == "red"

    def test_boundary_minus_5(self):
        status, color = get_status(-5)
        assert status == "lag_behind"
        assert color == "orange"

    def test_boundary_zero(self):
        status, color = get_status(0)
        assert status == "lag_on_track"
        assert color == "green"

    def test_boundary_5(self):
        status, color = get_status(5)
        assert status == "lag_ahead"
        assert color == "blue"

    def test_custom_translation_function(self):
        t_func = lambda x: x.upper()
        status, color = get_status(-15, t_func=t_func)
        assert status == "LAG_URGENT"


# ── calculate_nth_lag_lead ───────────────────────────────────────────────────


def _make_project_master(projects=None):
    """Helper: build a project_master DataFrame."""
    if projects is None:
        today = datetime.now()
        projects = [
            {
                "Project No": "C2264",
                "Title": "Test Project A",
                "Start Date": today - timedelta(days=180),
                "End Date": today + timedelta(days=180),
            },
            {
                "Project No": "C2265",
                "Title": "Test Project B",
                "Start Date": today - timedelta(days=90),
                "End Date": today + timedelta(days=270),
            },
        ]
    return pd.DataFrame(projects)


@pytest.mark.unit
class TestCalculateNthLagLead:

    def test_empty_project_master_returns_empty(self):
        result = calculate_nth_lag_lead(pd.DataFrame(), {}, {})
        assert result == []

    def test_none_project_master_returns_empty(self):
        result = calculate_nth_lag_lead(None, {}, {})
        assert result == []

    def test_valid_data_produces_correct_fields(self):
        pm = _make_project_master()
        config = {
            "C2264": {"target_qty": 1000, "productivity": 3.0, "skip": False},
            "C2265": {"target_qty": 500, "productivity": 2.5, "skip": False},
        }
        actual_qty = {"C2264": 400, "C2265": 100}

        results = calculate_nth_lag_lead(pm, config, actual_qty)

        assert len(results) == 2
        expected_keys = {
            "Project", "Title", "Start Date", "End Date",
            "Target Qty", "Actual Qty", "Target % (Linear)",
            "Progress %", "Productivity", "NTH Lag/Lead",
            "Status", "Status Color",
        }
        assert expected_keys == set(results[0].keys())

    def test_skips_projects_with_skip_true(self):
        pm = _make_project_master()
        config = {
            "C2264": {"target_qty": 1000, "productivity": 3.0, "skip": True},
            "C2265": {"target_qty": 500, "productivity": 2.5, "skip": False},
        }
        actual_qty = {"C2264": 400, "C2265": 100}

        results = calculate_nth_lag_lead(pm, config, actual_qty)
        project_ids = [r["Project"] for r in results]

        assert "C2264" not in project_ids
        assert "C2265" in project_ids

    def test_skips_projects_without_target_qty(self):
        pm = _make_project_master()
        config = {
            "C2264": {"target_qty": None, "productivity": 3.0, "skip": False},
        }
        results = calculate_nth_lag_lead(pm, config, {})
        assert len(results) == 0


# ── export_lag_report ────────────────────────────────────────────────────────


@pytest.mark.unit
class TestExportLagReport:

    def test_creates_excel_file(self, tmp_path):
        results_df = pd.DataFrame([
            {
                "Project": "C2264",
                "Title": "Test",
                "NTH Lag/Lead": -5.0,
                "Status": "lag_behind",
                "Status Color": "orange",
            }
        ])
        output = tmp_path / "lag_report.xlsx"
        success = export_lag_report(results_df, output)

        assert success is True
        assert output.exists()

    def test_removes_status_color_column(self, tmp_path):
        results_df = pd.DataFrame([
            {
                "Project": "C2264",
                "Status Color": "red",
                "Status": "lag_urgent",
            }
        ])
        output = tmp_path / "lag_report.xlsx"
        export_lag_report(results_df, output)

        exported = pd.read_excel(output, sheet_name="Lag Analysis")
        assert "Status Color" not in exported.columns


# ── LagAnalyzer ──────────────────────────────────────────────────────────────


@pytest.mark.unit
class TestLagAnalyzer:

    def test_calculate_returns_false_when_no_project_master(self):
        analyzer = LagAnalyzer()
        result = analyzer.calculate(actual_qty_map={"C2264": 100})
        assert result is False

    def test_export_returns_false_when_no_results(self, tmp_path):
        analyzer = LagAnalyzer()
        result = analyzer.export(tmp_path / "out.xlsx")
        assert result is False


# ── load_project_master ─────────────────────────────────────────────────────


def _create_project_master_excel(filepath, headers=None, rows=None):
    """Helper: create a project master Excel file using openpyxl."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active

    if headers is None:
        headers = ["ID", "Category", "Project No", "One Line Title",
                   "Start Date", "Finish Date", "Target Qty"]
    if rows is None:
        rows = [
            [1, "OHL", "C2264", "Test Project A", "01/01/2024", "31/12/2025", 1000],
            [2, "OHL", "C2265", "Test Project B", "01/06/2024", "30/06/2026", 500],
        ]

    for col_idx, header in enumerate(headers, 1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row_data in enumerate(rows, 2):
        for col_idx, val in enumerate(row_data, 1):
            ws.cell(row=row_idx, column=col_idx, value=val)

    wb.save(filepath)


@pytest.mark.unit
class TestLoadProjectMaster:

    def test_loads_project_master_from_excel(self, tmp_path):
        filepath = tmp_path / "master.xlsx"
        _create_project_master_excel(filepath)

        df, descriptions, target_map = load_project_master(filepath)

        assert not df.empty
        assert len(df) == 2
        assert "C2264" in df["Project No"].values
        assert "C2265" in df["Project No"].values

    def test_extracts_descriptions(self, tmp_path):
        filepath = tmp_path / "master.xlsx"
        _create_project_master_excel(filepath)

        _, descriptions, _ = load_project_master(filepath)

        assert descriptions["C2264"] == "Test Project A"
        assert descriptions["C2265"] == "Test Project B"

    def test_extracts_target_qty(self, tmp_path):
        filepath = tmp_path / "master.xlsx"
        _create_project_master_excel(filepath)

        _, _, target_map = load_project_master(filepath)

        assert target_map["C2264"] == 1000.0
        assert target_map["C2265"] == 500.0

    def test_case_insensitive_column_mapping(self, tmp_path):
        filepath = tmp_path / "master.xlsx"
        headers = ["ID", "Cat", "project no", "one line title",
                   "start date", "finish date", "target qty"]
        rows = [
            [1, "OHL", "C3000", "Lower Case Headers", "01/01/2024", "31/12/2025", 750],
        ]
        _create_project_master_excel(filepath, headers=headers, rows=rows)

        df, descriptions, target_map = load_project_master(filepath)

        assert "C3000" in df["Project No"].values
        assert descriptions["C3000"] == "Lower Case Headers"
        assert target_map["C3000"] == 750.0

    def test_fallback_to_column_c_when_no_project_no_header(self, tmp_path):
        filepath = tmp_path / "master.xlsx"
        # No column named "Project No" - should fall back to column C (index 2)
        headers = ["ID", "Category", "Code", "Title"]
        rows = [
            [1, "OHL", "C4000", "Fallback Project"],
        ]
        _create_project_master_excel(filepath, headers=headers, rows=rows)

        df, _, _ = load_project_master(filepath)

        assert "C4000" in df["Project No"].values

    def test_skips_empty_project_no_rows(self, tmp_path):
        filepath = tmp_path / "master.xlsx"
        headers = ["ID", "Category", "Project No", "One Line Title", "Target Qty"]
        rows = [
            [1, "OHL", "C2264", "Valid", 100],
            [2, "OHL", "", "Empty Code", 200],
            [3, "OHL", "C2265", "Also Valid", 300],
        ]
        _create_project_master_excel(filepath, headers=headers, rows=rows)

        df, _, _ = load_project_master(filepath)

        assert len(df) == 2
        assert "C2264" in df["Project No"].values
        assert "C2265" in df["Project No"].values


# ── LagAnalyzer.load_project_master ─────────────────────────────────────────


@pytest.mark.unit
class TestLagAnalyzerLoadProjectMaster:

    def test_load_project_master_success(self, tmp_path):
        filepath = tmp_path / "master.xlsx"
        _create_project_master_excel(filepath)

        analyzer = LagAnalyzer()
        result = analyzer.load_project_master(filepath)

        assert result is True
        assert analyzer.project_master is not None
        assert len(analyzer.projects) == 2
        assert "C2264" in analyzer.projects

    def test_load_project_master_initializes_config(self, tmp_path):
        filepath = tmp_path / "master.xlsx"
        _create_project_master_excel(filepath)

        analyzer = LagAnalyzer()
        analyzer.load_project_master(filepath)

        assert "C2264" in analyzer.config
        assert analyzer.config["C2264"]["productivity"] == 3.0
        assert analyzer.config["C2264"]["skip"] is False

    def test_load_project_master_invalid_file(self, tmp_path):
        filepath = tmp_path / "nonexistent.xlsx"
        analyzer = LagAnalyzer()
        result = analyzer.load_project_master(filepath)
        assert result is False


# ── LagAnalyzer._match_productivity ─────────────────────────────────────────


@pytest.mark.unit
class TestLagAnalyzerMatchProductivity:

    def test_match_productivity_from_summary(self, tmp_path):
        filepath = tmp_path / "master.xlsx"
        _create_project_master_excel(filepath)

        analyzer = LagAnalyzer()
        analyzer.load_project_master(filepath)

        summary_df = pd.DataFrame([
            {"Project": "C2264", "Qty per NTH": 4.5},
            {"Project": "C2265", "Qty per NTH": 2.8},
        ])
        analyzer._match_productivity(summary_df)

        assert analyzer.productivity_map["C2264"] == 4.5
        assert analyzer.productivity_map["C2265"] == 2.8
        assert analyzer.config["C2264"]["productivity"] == 4.5
        assert analyzer.config["C2264"]["source"] == "daily_report"

    def test_match_productivity_with_summary_on_load(self, tmp_path):
        filepath = tmp_path / "master.xlsx"
        _create_project_master_excel(filepath)

        summary_df = pd.DataFrame([
            {"Project": "C2264", "Qty per NTH": 5.0},
        ])

        analyzer = LagAnalyzer()
        analyzer.load_project_master(filepath, summary_df=summary_df)

        assert analyzer.productivity_map["C2264"] == 5.0


# ── LagAnalyzer.calculate ──────────────────────────────────────────────────


@pytest.mark.unit
class TestLagAnalyzerCalculate:

    def test_calculate_with_valid_data(self, tmp_path):
        filepath = tmp_path / "master.xlsx"
        _create_project_master_excel(filepath)

        analyzer = LagAnalyzer()
        analyzer.load_project_master(filepath)

        # Set target_qty in config (loaded from Excel)
        actual_qty_map = {"C2264": 400, "C2265": 100}
        result = analyzer.calculate(actual_qty_map)

        assert result is True
        assert analyzer.results is not None
        assert len(analyzer.results) > 0
        assert analyzer.last_calculated is not None

    def test_calculate_returns_false_with_no_matching_projects(self, tmp_path):
        filepath = tmp_path / "master.xlsx"
        # Projects with no target_qty in Excel
        headers = ["ID", "Category", "Project No", "One Line Title"]
        rows = [
            [1, "OHL", "C9999", "No Target"],
        ]
        _create_project_master_excel(filepath, headers=headers, rows=rows)

        analyzer = LagAnalyzer()
        analyzer.load_project_master(filepath)
        result = analyzer.calculate({"C9999": 100})

        # No target_qty => config has None => skipped => no results
        assert result is False


# ── LagAnalyzer.export ──────────────────────────────────────────────────────


@pytest.mark.unit
class TestLagAnalyzerExport:

    def test_export_creates_excel(self, tmp_path):
        filepath = tmp_path / "master.xlsx"
        _create_project_master_excel(filepath)

        analyzer = LagAnalyzer()
        analyzer.load_project_master(filepath)
        analyzer.calculate({"C2264": 400, "C2265": 100})

        output = tmp_path / "lag_export.xlsx"
        result = analyzer.export(output)

        assert result is True
        assert output.exists()

        # Verify content
        exported = pd.read_excel(output, sheet_name="Lag Analysis")
        assert "Project" in exported.columns
        assert "Status Color" not in exported.columns
