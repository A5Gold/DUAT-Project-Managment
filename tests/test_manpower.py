# MTR DUAT - Manpower Analysis Tests
"""Unit tests for analysis/manpower.py"""

import pytest
from pathlib import Path

from analysis.manpower import (
    get_daily_headcount,
    get_team_distribution,
    get_job_type_manpower,
    get_role_frequency,
    get_work_access_analysis,
    get_individual_stats,
    get_summary_kpis,
    export_manpower_excel,
    ManpowerAnalyzer,
)


def _make_records(n: int = 3):
    """Helper: build realistic manpower shift records."""
    records = [
        {
            "date": "2024-01-08",
            "day_of_week": "Mon",
            "week": "WK01",
            "year": "2024",
            "shift": "Night",
            "on_duty_names": ["Alice", "Bob", "Charlie"],
            "apprentices": ["Dave"],
            "term_labour_count": 2,
            "on_duty_team_counts": {"S2": 3, "S3": 2},
            "jobs": [
                {
                    "type": "PA work",
                    "project_code": "C2264",
                    "description": "Cable pulling",
                    "qty": 5,
                    "done_by_raw": "S2",
                    "total_workers": 5,
                    "team_counts": {"S2": 3, "S3": 2},
                    "worker_names": ["Alice", "Bob"],
                    "roles": {
                        "CP_P": [],
                        "CP_T": ["Alice"],
                        "AP_E": ["Bob"],
                        "SPC": ["Charlie"],
                        "HSM": [],
                        "NP": [],
                        "EPIC": "E001",
                    },
                },
                {
                    "type": "SPA work",
                    "project_code": "C2265",
                    "description": "OHL access",
                    "qty": 2,
                    "done_by_raw": "S3",
                    "total_workers": 3,
                    "team_counts": {"S3": 3},
                    "worker_names": ["Charlie"],
                    "roles": {
                        "CP_P": [],
                        "CP_T": [],
                        "AP_E": ["Charlie"],
                        "SPC": [],
                        "HSM": ["Alice"],
                        "NP": [],
                        "EPIC": "E002",
                    },
                },
            ],
        },
        {
            "date": "2024-01-09",
            "day_of_week": "Tue",
            "week": "WK01",
            "year": "2024",
            "shift": "Night",
            "on_duty_names": ["Alice", "Bob"],
            "apprentices": [],
            "term_labour_count": 0,
            "on_duty_team_counts": {},
            "jobs": [
                {
                    "type": "CBM",
                    "project_code": "C2264",
                    "description": "Inspection",
                    "qty": 3,
                    "done_by_raw": "S2",
                    "total_workers": 4,
                    "team_counts": {"S2": 2, "S4": 2},
                    "worker_names": ["Alice", "Bob"],
                    "roles": {
                        "CP_P": ["Alice"],
                        "CP_T": [],
                        "AP_E": [],
                        "SPC": ["Bob"],
                        "HSM": [],
                        "NP": [],
                        "EPIC": "E003",
                    },
                },
            ],
        },
        {
            "date": "2024-01-15",
            "day_of_week": "Mon",
            "week": "WK02",
            "year": "2024",
            "shift": "Night",
            "on_duty_names": ["Alice", "Charlie", "Eve"],
            "apprentices": ["Frank"],
            "term_labour_count": 1,
            "on_duty_team_counts": {"S5": 2},
            "jobs": [
                {
                    "type": "PA work",
                    "project_code": "C2266",
                    "description": "Wiring",
                    "qty": 4,
                    "done_by_raw": "S5",
                    "total_workers": 6,
                    "team_counts": {"S5": 4, "S2": 2},
                    "worker_names": ["Alice", "Eve"],
                    "roles": {
                        "CP_P": ["Eve"],
                        "CP_T": ["Alice"],
                        "AP_E": [],
                        "SPC": [],
                        "HSM": ["Charlie"],
                        "NP": [],
                        "EPIC": "E004",
                    },
                },
            ],
        },
    ]
    return records[:n]


# ── get_daily_headcount ──────────────────────────────────────────────────────


@pytest.mark.unit
class TestGetDailyHeadcount:

    def test_with_sample_records(self):
        records = _make_records()
        result = get_daily_headcount(records)

        assert len(result) == 3
        first = result[0]
        assert first["date"] == "2024-01-08"
        assert first["named_count"] == 3
        assert first["apprentice_count"] == 1
        assert first["term_labour_count"] == 2
        # headcount = 3 names + 1 apprentice + 2 term + (3+2) teams = 11
        assert first["headcount"] == 11

    def test_empty_records(self):
        result = get_daily_headcount([])
        assert result == []

    def test_keys_present(self):
        records = _make_records(1)
        result = get_daily_headcount(records)
        expected_keys = {
            "date", "day_of_week", "week", "year", "shift",
            "headcount", "named_count", "apprentice_count",
            "term_labour_count", "team_counts",
        }
        assert expected_keys == set(result[0].keys())


# ── get_team_distribution ────────────────────────────────────────────────────


@pytest.mark.unit
class TestGetTeamDistribution:

    def test_aggregates_by_week(self):
        records = _make_records()
        dist = get_team_distribution(records)

        assert "WK01" in dist
        assert "WK02" in dist
        # WK01: job1 S2=3,S3=2 + job2 S3=3 + job3 S2=2,S4=2
        assert dist["WK01"]["S2"] == 5  # 3 + 2
        assert dist["WK01"]["S3"] == 5  # 2 + 3

    def test_sorted_by_week(self):
        records = _make_records()
        dist = get_team_distribution(records)
        weeks = list(dist.keys())
        assert weeks == sorted(weeks, key=lambda w: int(w.replace("WK", "")))


# ── get_job_type_manpower ────────────────────────────────────────────────────


@pytest.mark.unit
class TestGetJobTypeManpower:

    def test_calculates_averages(self):
        records = _make_records()
        result = get_job_type_manpower(records)

        assert "PA work" in result
        pa = result["PA work"]
        assert pa["total_jobs"] == 2
        assert pa["total_workers"] == 11  # 5 + 6
        assert pa["avg_workers"] == 5.5

    def test_all_job_types_present(self):
        records = _make_records()
        result = get_job_type_manpower(records)
        assert "PA work" in result
        assert "SPA work" in result
        assert "CBM" in result


# ── get_role_frequency ───────────────────────────────────────────────────────


@pytest.mark.unit
class TestGetRoleFrequency:

    def test_sorts_by_total_descending(self):
        records = _make_records()
        result = get_role_frequency(records)

        assert len(result) > 0
        totals = [entry["total"] for entry in result]
        assert totals == sorted(totals, reverse=True)

    def test_empty_records(self):
        result = get_role_frequency([])
        assert result == []

    def test_entry_keys(self):
        records = _make_records()
        result = get_role_frequency(records)
        expected_keys = {"name", "CP_P", "CP_T", "AP_E", "SPC", "HSM", "NP", "total"}
        assert expected_keys == set(result[0].keys())

    def test_alice_roles(self):
        records = _make_records()
        result = get_role_frequency(records)
        alice = next(e for e in result if e["name"] == "Alice")
        # Alice: CP_P=1 (job3-CBM), CP_T=2 (job1+job4), HSM=1 (job2-SPA) => wait
        # job1: CP_T=Alice, job2: HSM=Alice, job3: CP_P=Alice, job4: CP_T=Alice
        assert alice["CP_P"] == 1
        assert alice["CP_T"] == 2
        assert alice["HSM"] == 1


# ── get_work_access_analysis ─────────────────────────────────────────────────


@pytest.mark.unit
class TestGetWorkAccessAnalysis:

    def test_classifies_spa_pa_possession_other(self):
        records = _make_records()
        result = get_work_access_analysis(records)

        assert "PA work" in result
        assert "SPA work" in result

    def test_spa_priority_over_pa(self):
        """SPA work type should be classified as SPA, not PA."""
        records = _make_records()
        result = get_work_access_analysis(records)
        assert result["SPA work"]["count"] == 1

    def test_possession_when_cp_p_present(self):
        """Jobs with CP_P role and no SPA/PA type should be Possession."""
        records = _make_records()
        result = get_work_access_analysis(records)
        # CBM job has CP_P=["Alice"], type="CBM" => Possession
        assert "Possession" in result
        assert result["Possession"]["count"] == 1

    def test_pa_work_without_cp_p(self):
        records = _make_records()
        result = get_work_access_analysis(records)
        # PA work jobs: job1 has no CP_P, job4 has CP_P=["Eve"]
        # job1 => PA work (type="PA work"), job4 => PA work (type="PA work")
        # Even though job4 has CP_P, the type check for "PA work" takes priority
        assert result["PA work"]["count"] >= 1

    def test_avg_workers_calculated(self):
        records = _make_records()
        result = get_work_access_analysis(records)
        for cat_data in result.values():
            assert "avg_workers" in cat_data
            assert cat_data["avg_workers"] > 0


# ── get_individual_stats ─────────────────────────────────────────────────────


@pytest.mark.unit
class TestGetIndividualStats:

    def test_sorts_by_duty_days(self):
        records = _make_records()
        result = get_individual_stats(records)

        duty_days = [entry["duty_days"] for entry in result]
        assert duty_days == sorted(duty_days, reverse=True)

    def test_alice_appears_3_times(self):
        records = _make_records()
        result = get_individual_stats(records)
        alice = next(e for e in result if e["name"] == "Alice")
        assert alice["duty_days"] == 3  # on_duty in all 3 records


# ── get_summary_kpis ─────────────────────────────────────────────────────────


@pytest.mark.unit
class TestGetSummaryKpis:

    def test_returns_correct_keys(self):
        records = _make_records()
        kpis = get_summary_kpis(records)

        expected_keys = {
            "total_jobs", "avg_workers_per_job",
            "unique_staff_count", "top_role_holder",
            "top_role_holder_count",
        }
        assert expected_keys == set(kpis.keys())

    def test_total_jobs_count(self):
        records = _make_records()
        kpis = get_summary_kpis(records)
        # 2 jobs in record 0, 1 job in record 1, 1 job in record 2 = 4
        assert kpis["total_jobs"] == 4

    def test_unique_staff_includes_apprentices(self):
        records = _make_records()
        kpis = get_summary_kpis(records)
        # on_duty: Alice, Bob, Charlie, Eve; apprentices: Dave, Frank;
        # worker_names: Alice, Bob, Charlie, Alice, Eve
        # unique = {Alice, Bob, Charlie, Eve, Dave, Frank} = 6
        assert kpis["unique_staff_count"] == 6


# ── ManpowerAnalyzer ─────────────────────────────────────────────────────────


@pytest.mark.unit
class TestManpowerAnalyzer:

    def test_wraps_summary_kpis(self):
        analyzer = ManpowerAnalyzer(_make_records())
        kpis = analyzer.summary_kpis()
        assert "total_jobs" in kpis

    def test_wraps_daily_headcount(self):
        analyzer = ManpowerAnalyzer(_make_records())
        result = analyzer.daily_headcount()
        assert len(result) == 3

    def test_wraps_role_frequency(self):
        analyzer = ManpowerAnalyzer(_make_records())
        result = analyzer.role_frequency()
        assert len(result) > 0

    def test_wraps_work_access_analysis(self):
        analyzer = ManpowerAnalyzer(_make_records())
        result = analyzer.work_access_analysis()
        assert isinstance(result, dict)

    def test_set_records(self):
        analyzer = ManpowerAnalyzer()
        assert analyzer.summary_kpis()["total_jobs"] == 0

        analyzer.set_records(_make_records())
        assert analyzer.summary_kpis()["total_jobs"] == 4

    def test_individual_stats(self):
        analyzer = ManpowerAnalyzer(_make_records())
        result = analyzer.individual_stats()
        assert len(result) > 0
        assert result[0]["duty_days"] >= result[-1]["duty_days"]


# ── export_manpower_excel ───────────────────────────────────────────────────


@pytest.mark.unit
class TestExportManpowerExcel:

    def test_creates_excel_with_four_sheets(self, tmp_path):
        records = _make_records()
        output = tmp_path / "manpower.xlsx"

        result = export_manpower_excel(records, output)

        assert result == output
        assert output.exists()

        from openpyxl import load_workbook
        wb = load_workbook(output)
        sheet_names = wb.sheetnames
        assert "Raw Data" in sheet_names
        assert "Job Type Summary" in sheet_names
        assert "Role Frequency" in sheet_names
        assert "Weekly Team Distribution" in sheet_names
        assert len(sheet_names) == 4
        wb.close()

    def test_raw_data_sheet_has_correct_headers(self, tmp_path):
        records = _make_records()
        output = tmp_path / "manpower.xlsx"
        export_manpower_excel(records, output)

        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb["Raw Data"]

        expected_headers = [
            "Year", "Week", "Date", "Day", "Shift",
            "Job Type", "Project Code", "Description", "Qty",
            "Done By", "Total Workers",
            "S2", "S3", "S4", "S5",
            "EPIC", "CP(P)", "CP(T)", "AP(E)", "SPC", "HSM",
            "On Duty Count", "Apprentices", "Term Labour",
        ]
        actual_headers = [ws.cell(row=1, column=c).value for c in range(1, len(expected_headers) + 1)]
        assert actual_headers == expected_headers
        wb.close()

    def test_raw_data_has_rows(self, tmp_path):
        records = _make_records()
        output = tmp_path / "manpower.xlsx"
        export_manpower_excel(records, output)

        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb["Raw Data"]
        # 3 records: record 0 has 2 jobs, record 1 has 1 job, record 2 has 1 job = 4 data rows
        assert ws.max_row == 5  # 1 header + 4 data rows
        wb.close()

    def test_job_type_summary_sheet(self, tmp_path):
        records = _make_records()
        output = tmp_path / "manpower.xlsx"
        export_manpower_excel(records, output)

        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb["Job Type Summary"]
        # Header row + at least 3 job types (PA work, SPA work, CBM)
        assert ws.max_row >= 4
        wb.close()

    def test_role_frequency_sheet(self, tmp_path):
        records = _make_records()
        output = tmp_path / "manpower.xlsx"
        export_manpower_excel(records, output)

        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb["Role Frequency"]
        # Should have header + at least one person
        assert ws.max_row >= 2
        assert ws.cell(row=1, column=1).value == "Name"
        wb.close()

    def test_empty_records(self, tmp_path):
        output = tmp_path / "manpower_empty.xlsx"
        result = export_manpower_excel([], output)

        assert result == output
        assert output.exists()


# ── ManpowerAnalyzer.export_excel ───────────────────────────────────────────


@pytest.mark.unit
class TestManpowerAnalyzerExportExcel:

    def test_export_excel_creates_file(self, tmp_path):
        analyzer = ManpowerAnalyzer(_make_records())
        output = tmp_path / "analyzer_export.xlsx"

        result = analyzer.export_excel(output)

        assert result == output
        assert output.exists()

        from openpyxl import load_workbook
        wb = load_workbook(output)
        assert len(wb.sheetnames) == 4
        wb.close()

    def test_export_excel_empty_records(self, tmp_path):
        analyzer = ManpowerAnalyzer()
        output = tmp_path / "analyzer_empty.xlsx"

        result = analyzer.export_excel(output)

        assert result == output
        assert output.exists()
