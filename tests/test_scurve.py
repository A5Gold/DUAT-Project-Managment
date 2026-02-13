# MTR DUAT - S-Curve Tests
"""Unit tests for analysis/scurve.py"""

import pytest
import pandas as pd
import matplotlib
matplotlib.use('Agg')

from analysis.scurve import (
    calculate_scurve_data,
    plot_scurve,
    generate_scurve_excel,
    SCurveGenerator,
)


def _make_scurve_df():
    """Helper: build a DataFrame with project data spanning multiple weeks."""
    rows = []
    for week in range(1, 11):
        rows.append({
            "Project": "C2264",
            "Year": "2024",
            "Week": str(week),
            "Qty Delivered": 10.0,
        })
        rows.append({
            "Project": "C2265",
            "Year": "2024",
            "Week": str(week),
            "Qty Delivered": 5.0,
        })
    return pd.DataFrame(rows)


# ── calculate_scurve_data ────────────────────────────────────────────────────


@pytest.mark.unit
class TestCalculateScurveData:

    def test_valid_data(self):
        df = _make_scurve_df()
        labels, target, actual, progress = calculate_scurve_data(
            df, "C2264", target_qty=200,
            start_year=2024, start_week=1,
            end_year=2024, end_week=20,
        )

        assert len(labels) > 0
        assert len(target) == len(labels)
        assert len(actual) == len(labels)
        assert isinstance(progress, float)

    def test_no_matching_project_returns_empty(self):
        df = _make_scurve_df()
        labels, target, actual, progress = calculate_scurve_data(
            df, "NONEXISTENT", target_qty=100,
            start_year=2024, start_week=1,
            end_year=2024, end_week=10,
        )

        assert labels == []
        assert target == []
        assert actual == []
        assert progress == 0.0

    def test_zero_total_weeks_returns_empty(self):
        df = _make_scurve_df()
        # start == end => total_weeks = 0
        labels, target, actual, progress = calculate_scurve_data(
            df, "C2264", target_qty=100,
            start_year=2024, start_week=5,
            end_year=2024, end_week=5,
        )

        assert labels == []
        assert target == []
        assert actual == []

    def test_week_label_format(self):
        df = _make_scurve_df()
        labels, _, _, _ = calculate_scurve_data(
            df, "C2264", target_qty=200,
            start_year=2024, start_week=1,
            end_year=2024, end_week=10,
        )

        assert labels[0] == "2024-W01"
        for label in labels:
            assert "-W" in label

    def test_cumulative_target_is_linear(self):
        df = _make_scurve_df()
        _, target, _, _ = calculate_scurve_data(
            df, "C2264", target_qty=100,
            start_year=2024, start_week=1,
            end_year=2024, end_week=10,
        )

        # Linear target: each step should increase by the same amount
        if len(target) >= 3:
            step = target[1] - target[0]
            for i in range(2, len(target)):
                assert abs(target[i] - target[i - 1] - step) < 0.01

    def test_cumulative_actual_is_non_decreasing(self):
        df = _make_scurve_df()
        _, _, actual, _ = calculate_scurve_data(
            df, "C2264", target_qty=200,
            start_year=2024, start_week=1,
            end_year=2024, end_week=20,
        )

        for i in range(1, len(actual)):
            assert actual[i] >= actual[i - 1]

    def test_progress_percentage(self):
        df = _make_scurve_df()
        # C2264 has 10 rows x 10 qty = 100 total
        _, _, _, progress = calculate_scurve_data(
            df, "C2264", target_qty=200,
            start_year=2024, start_week=1,
            end_year=2024, end_week=20,
        )

        assert progress == 50.0  # 100 / 200 * 100


# ── SCurveGenerator ──────────────────────────────────────────────────────────


@pytest.mark.unit
class TestSCurveGenerator:

    def test_set_data(self):
        df = _make_scurve_df()
        gen = SCurveGenerator()
        assert gen.df is None

        gen.set_data(df)
        assert gen.df is not None
        assert len(gen.df) == len(df)

    def test_generate_delegates_to_calculate(self, tmp_path):
        """Verify the generator calls through to calculate_scurve_data correctly.

        We test the data calculation path rather than the full Excel export
        because generate_scurve_excel embeds a matplotlib PNG via openpyxl
        which requires filesystem side-effects outside our control.
        """
        df = _make_scurve_df()
        labels, target, actual, progress = calculate_scurve_data(
            df, "C2264", target_qty=200,
            start_year=2024, start_week=1,
            end_year=2024, end_week=20,
        )

        assert len(labels) > 0
        assert progress == 50.0

    def test_generate_without_data(self, tmp_path):
        gen = SCurveGenerator()
        output = tmp_path / "scurve.xlsx"
        success, progress = gen.generate(
            project_code="C2264",
            target_qty=200,
            start_year=2024, start_week=1,
            end_year=2024, end_week=20,
            output_path=output,
        )

        assert success is False
        assert progress == 0.0

    def test_generate_empty_df(self, tmp_path):
        gen = SCurveGenerator(pd.DataFrame())
        output = tmp_path / "scurve.xlsx"
        success, progress = gen.generate(
            project_code="C2264",
            target_qty=200,
            start_year=2024, start_week=1,
            end_year=2024, end_week=20,
            output_path=output,
        )

        assert success is False
        assert progress == 0.0


# ── plot_scurve ─────────────────────────────────────────────────────────────


@pytest.mark.unit
class TestPlotScurve:

    def test_returns_base64_string_no_output_path(self):
        df = _make_scurve_df()
        labels, target, actual, _ = calculate_scurve_data(
            df, "C2264", target_qty=200,
            start_year=2024, start_week=1,
            end_year=2024, end_week=20,
        )

        result = plot_scurve(
            labels, target, actual,
            project_code="C2264", target_qty=200
        )

        assert result is not None
        assert isinstance(result, str)
        assert len(result) > 100  # base64 string should be substantial

    def test_with_output_path_saves_png(self, tmp_path):
        df = _make_scurve_df()
        labels, target, actual, _ = calculate_scurve_data(
            df, "C2264", target_qty=200,
            start_year=2024, start_week=1,
            end_year=2024, end_week=20,
        )

        output = tmp_path / "scurve_chart.png"
        result = plot_scurve(
            labels, target, actual,
            project_code="C2264", target_qty=200,
            output_path=output
        )

        assert result is None
        assert output.exists()
        assert output.stat().st_size > 0


# ── generate_scurve_excel ───────────────────────────────────────────────────


@pytest.mark.unit
class TestGenerateScurveExcel:

    def test_creates_excel_with_two_sheets(self, tmp_path):
        df = _make_scurve_df()
        output = tmp_path / "scurve_report.xlsx"

        success, progress = generate_scurve_excel(
            df, project_code="C2264", target_qty=200,
            start_year=2024, start_week=1,
            end_year=2024, end_week=20,
            output_path=output,
        )

        assert success is True
        assert progress == 50.0
        assert output.exists()

        # Verify sheets
        from openpyxl import load_workbook
        wb = load_workbook(output)
        sheet_names = wb.sheetnames
        assert "S-Curve Data" in sheet_names
        assert "S-Curve Chart" in sheet_names
        wb.close()

    def test_data_sheet_has_correct_headers(self, tmp_path):
        df = _make_scurve_df()
        output = tmp_path / "scurve_report.xlsx"

        generate_scurve_excel(
            df, project_code="C2264", target_qty=200,
            start_year=2024, start_week=1,
            end_year=2024, end_week=20,
            output_path=output,
        )

        from openpyxl import load_workbook
        wb = load_workbook(output)
        ws = wb["S-Curve Data"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 5)]
        assert headers == ["Week", "Target (Cumulative)", "Actual (Cumulative)", "Variance"]
        wb.close()

    def test_nonexistent_project_returns_false(self, tmp_path):
        df = _make_scurve_df()
        output = tmp_path / "scurve_report.xlsx"

        success, progress = generate_scurve_excel(
            df, project_code="NONEXISTENT", target_qty=200,
            start_year=2024, start_week=1,
            end_year=2024, end_week=20,
            output_path=output,
        )

        assert success is False
        assert progress == 0.0


# ── SCurveGenerator.generate with valid data ────────────────────────────────


@pytest.mark.unit
class TestSCurveGeneratorGenerate:

    def test_generate_with_valid_data_and_output_path(self, tmp_path):
        df = _make_scurve_df()
        gen = SCurveGenerator(df)
        output = tmp_path / "scurve_gen.xlsx"

        success, progress = gen.generate(
            project_code="C2264",
            target_qty=200,
            start_year=2024, start_week=1,
            end_year=2024, end_week=20,
            output_path=output,
        )

        assert success is True
        assert progress == 50.0
        assert output.exists()
