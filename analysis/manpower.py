# MTR PS-OHLR DUAT - Manpower Analysis
"""
Aggregate manpower records and generate analysis outputs.
"""

import logging
from collections import defaultdict
from typing import List, Dict, Any, Tuple
from pathlib import Path
from datetime import datetime

logger = logging.getLogger(__name__)

# ── Aggregation Functions ─────────────────────────────────────────────────────


def get_daily_headcount(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Calculate headcount per date and shift.

    Returns list of dicts with keys:
        date, day_of_week, week, shift, headcount, team_counts
    """
    results = []
    for record in records:
        names = record.get("on_duty_names", [])
        apprentices = record.get("apprentices", [])
        term_count = record.get("term_labour_count", 0)
        team_counts = record.get("on_duty_team_counts", {})
        team_total = sum(team_counts.values())

        headcount = len(names) + len(apprentices) + term_count + team_total

        results.append({
            "date": record["date"],
            "day_of_week": record["day_of_week"],
            "week": record["week"],
            "year": record["year"],
            "shift": record["shift"],
            "headcount": headcount,
            "named_count": len(names),
            "apprentice_count": len(apprentices),
            "term_labour_count": term_count,
            "team_counts": dict(team_counts),
        })
    return results


def get_team_distribution(records: List[Dict[str, Any]]) -> Dict[str, Dict[str, int]]:
    """
    Calculate team (S2/S3/S4/S5) allocation aggregated per week.

    Returns dict: {week: {S2: total, S3: total, S4: total, S5: total}}
    """
    weekly_teams = defaultdict(lambda: defaultdict(int))

    for record in records:
        week = record["week"]
        # Aggregate from job-level team counts
        for job in record.get("jobs", []):
            for team_id, count in job.get("team_counts", {}).items():
                weekly_teams[week][team_id] += count

    # Sort by week
    sorted_weeks = sorted(weekly_teams.keys(), key=lambda w: int(w.replace("WK", "")))
    return {week: dict(weekly_teams[week]) for week in sorted_weeks}


def get_job_type_manpower(records: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """
    Calculate average workers and role counts per job type.

    Returns dict: {
        job_type: {
            total_jobs: int,
            total_workers: int,
            avg_workers: float,
            avg_team_counts: {S2: float, S3: float, ...},
            avg_roles: {CP_P: float, CP_T: float, AP_E: float, SPC: float, HSM: float}
        }
    }
    """
    job_type_data = defaultdict(lambda: {
        "total_jobs": 0,
        "total_workers": 0,
        "team_sums": defaultdict(int),
        "role_sums": defaultdict(int),
    })

    for record in records:
        for job in record.get("jobs", []):
            jtype = job["type"]
            data = job_type_data[jtype]
            data["total_jobs"] += 1
            data["total_workers"] += job.get("total_workers", 0)

            for team_id, count in job.get("team_counts", {}).items():
                data["team_sums"][team_id] += count

            roles = job.get("roles", {})
            for role_key in ("CP_P", "CP_T", "AP_E", "SPC", "HSM", "NP"):
                role_count = len(roles.get(role_key, []))
                data["role_sums"][role_key] += role_count

    results = {}
    for jtype, data in job_type_data.items():
        total = data["total_jobs"]
        if total == 0:
            continue
        results[jtype] = {
            "total_jobs": total,
            "total_workers": data["total_workers"],
            "avg_workers": round(data["total_workers"] / total, 1),
            "avg_team_counts": {
                tid: round(cnt / total, 1)
                for tid, cnt in sorted(data["team_sums"].items())
            },
            "avg_roles": {
                role: round(cnt / total, 1)
                for role, cnt in sorted(data["role_sums"].items())
            },
        }

    return results


def get_role_frequency(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Calculate how often each person fills each EPIC role across all jobs.

    Returns list of dicts sorted by total assignments (descending):
        [{name, CP_P, CP_T, AP_E, SPC, HSM, NP, total}, ...]
    """
    person_roles = defaultdict(lambda: defaultdict(int))

    for record in records:
        for job in record.get("jobs", []):
            roles = job.get("roles", {})
            for role_key in ("CP_P", "CP_T", "AP_E", "SPC", "HSM", "NP"):
                for name in roles.get(role_key, []):
                    if name:
                        person_roles[name][role_key] += 1

    results = []
    for name, role_counts in person_roles.items():
        entry = {"name": name}
        total = 0
        for role_key in ("CP_P", "CP_T", "AP_E", "SPC", "HSM", "NP"):
            count = role_counts.get(role_key, 0)
            entry[role_key] = count
            total += count
        entry["total"] = total
        results.append(entry)

    results.sort(key=lambda x: x["total"], reverse=True)
    return results


def get_work_access_analysis(records: List[Dict[str, Any]]) -> Dict[str, Dict[str, Any]]:
    """
    Analyse jobs by work access type:
      - Possession: Engineering train in electrically isolated zone, identified by CP(P) role.
      - PA work:    Pedestrian Access, no engineering trains; isolation optional.
      - SPA work:   Special Pedestrian Access, may involve OHL access.
      - Other:      Jobs not matching above categories.

    Returns dict keyed by access type with counts and averages.
    """
    categories = {
        "Possession": {"count": 0, "workers": 0, "ap_count": 0, "spc_count": 0, "hsm_count": 0},
        "PA work": {"count": 0, "workers": 0, "ap_count": 0, "spc_count": 0, "hsm_count": 0},
        "SPA work": {"count": 0, "workers": 0, "ap_count": 0, "spc_count": 0, "hsm_count": 0},
        "Other": {"count": 0, "workers": 0, "ap_count": 0, "spc_count": 0, "hsm_count": 0},
    }

    for record in records:
        for job in record.get("jobs", []):
            roles = job.get("roles", {})
            job_type = job.get("type", "")
            has_cp_p = len(roles.get("CP_P", [])) > 0

            # Classify by access type
            if job_type == "SPA work":
                category = "SPA work"
            elif job_type == "PA work":
                category = "PA work"
            elif has_cp_p:
                # Possession work: CP(P) present means engineering train in isolated zone
                category = "Possession"
            else:
                category = "Other"

            cat = categories[category]
            cat["count"] += 1
            cat["workers"] += job.get("total_workers", 0)
            cat["ap_count"] += len(roles.get("AP_E", []))
            cat["spc_count"] += len(roles.get("SPC", []))
            cat["hsm_count"] += len(roles.get("HSM", []))

    results = {}
    for cat_name, cat_data in categories.items():
        count = cat_data["count"]
        if count == 0:
            continue
        results[cat_name] = {
            "count": count,
            "avg_workers": round(cat_data["workers"] / count, 1),
            "avg_aps": round(cat_data["ap_count"] / count, 1),
            "avg_spcs": round(cat_data["spc_count"] / count, 1),
            "avg_hsms": round(cat_data["hsm_count"] / count, 1),
        }

    return results


def get_individual_stats(records: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
    """
    Per-person on-duty count and role assignments.

    Returns list sorted by duty_days descending.
    """
    person_data = defaultdict(lambda: {"duty_days": 0, "roles_total": 0})

    for record in records:
        for name in record.get("on_duty_names", []):
            if name:
                person_data[name]["duty_days"] += 1

        for job in record.get("jobs", []):
            roles = job.get("roles", {})
            for role_key in ("CP_P", "CP_T", "AP_E", "SPC", "HSM", "NP"):
                for name in roles.get(role_key, []):
                    if name:
                        person_data[name]["roles_total"] += 1

    results = [
        {"name": name, "duty_days": data["duty_days"], "roles_total": data["roles_total"]}
        for name, data in person_data.items()
    ]
    results.sort(key=lambda x: x["duty_days"], reverse=True)
    return results


def get_summary_kpis(records: List[Dict[str, Any]]) -> Dict[str, Any]:
    """
    Calculate top-level KPI summary values.

    Returns dict with:
        total_jobs, avg_workers_per_job, unique_staff_count, top_role_holder
    """
    all_jobs = [job for r in records for job in r.get("jobs", [])]
    total_jobs = len(all_jobs)
    total_workers = sum(j.get("total_workers", 0) for j in all_jobs)

    # Unique staff from on-duty lists
    all_names = set()
    for record in records:
        for name in record.get("on_duty_names", []):
            if name:
                all_names.add(name)
        for name in record.get("apprentices", []):
            if name:
                all_names.add(name)
        for job in record.get("jobs", []):
            for name in job.get("worker_names", []):
                if name:
                    all_names.add(name)

    # Top role holder
    role_freq = get_role_frequency(records)
    top_holder = role_freq[0]["name"] if role_freq else "N/A"
    top_holder_count = role_freq[0]["total"] if role_freq else 0

    return {
        "total_jobs": total_jobs,
        "avg_workers_per_job": round(total_workers / total_jobs, 1) if total_jobs else 0,
        "unique_staff_count": len(all_names),
        "top_role_holder": top_holder,
        "top_role_holder_count": top_holder_count,
    }


# ── Excel Export ──────────────────────────────────────────────────────────────


def export_manpower_excel(records: List[Dict[str, Any]], filepath: Path) -> Path:
    """
    Export manpower analysis to a formatted Excel workbook.

    Sheets:
        1. Raw Data - All shift/job records flattened
        2. Job Type Summary - Avg workers and role breakdown per job type
        3. Role Frequency - Person x Role count matrix
        4. Weekly Team Distribution - S2/S3/S4/S5 per week

    Args:
        records: List of parsed shift records.
        filepath: Output Excel file path.

    Returns:
        Path to the saved file.
    """
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    workbook = Workbook()

    # Shared styles
    header_font = Font(name="Arial", size=10, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell_font = Font(name="Arial", size=9)
    thin_border = Border(
        left=Side(style="thin", color="CCCCCC"),
        right=Side(style="thin", color="CCCCCC"),
        top=Side(style="thin", color="CCCCCC"),
        bottom=Side(style="thin", color="CCCCCC"),
    )

    def style_header_row(sheet, num_cols):
        for col_idx in range(1, num_cols + 1):
            cell = sheet.cell(row=1, column=col_idx)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def style_data_cell(cell):
        cell.font = cell_font
        cell.border = thin_border
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    # ── Sheet 1: Raw Data ─────────────────────────────────────────────────
    ws_raw = workbook.active
    ws_raw.title = "Raw Data"

    raw_headers = [
        "Year", "Week", "Date", "Day", "Shift",
        "Job Type", "Project Code", "Description", "Qty",
        "Done By", "Total Workers",
        "S2", "S3", "S4", "S5",
        "EPIC", "CP(P)", "CP(T)", "AP(E)", "SPC", "HSM",
        "On Duty Count", "Apprentices", "Term Labour",
    ]
    for col_idx, header in enumerate(raw_headers, 1):
        ws_raw.cell(row=1, column=col_idx, value=header)
    style_header_row(ws_raw, len(raw_headers))

    row_num = 2
    for record in records:
        jobs = record.get("jobs", [])
        if not jobs:
            # Still write a row for shifts with no jobs
            jobs = [None]

        for job in jobs:
            values = [
                record["year"],
                record["week"],
                record["date"],
                record["day_of_week"],
                record["shift"],
            ]
            if job:
                roles = job.get("roles", {})
                team_c = job.get("team_counts", {})
                values.extend([
                    job.get("type", ""),
                    job.get("project_code", ""),
                    job.get("description", ""),
                    job.get("qty", 0),
                    job.get("done_by_raw", ""),
                    job.get("total_workers", 0),
                    team_c.get("S2", 0),
                    team_c.get("S3", 0),
                    team_c.get("S4", 0),
                    team_c.get("S5", 0),
                    roles.get("EPIC", ""),
                    ", ".join(roles.get("CP_P", [])),
                    ", ".join(roles.get("CP_T", [])),
                    ", ".join(roles.get("AP_E", [])),
                    ", ".join(roles.get("SPC", [])),
                    ", ".join(roles.get("HSM", [])),
                ])
            else:
                values.extend(["", "", "", 0, "", 0, 0, 0, 0, 0, "", "", "", "", "", ""])

            values.extend([
                len(record.get("on_duty_names", [])),
                ", ".join(record.get("apprentices", [])),
                record.get("term_labour_count", 0),
            ])

            for col_idx, val in enumerate(values, 1):
                cell = ws_raw.cell(row=row_num, column=col_idx, value=val)
                style_data_cell(cell)
            row_num += 1

    # Auto-fit column widths (approximate)
    for col_idx in range(1, len(raw_headers) + 1):
        ws_raw.column_dimensions[chr(64 + min(col_idx, 26))].width = 14

    # ── Sheet 2: Job Type Summary ─────────────────────────────────────────
    ws_summary = workbook.create_sheet("Job Type Summary")

    summary_headers = [
        "Job Type", "Total Jobs", "Total Workers", "Avg Workers/Job",
        "Avg S2", "Avg S3", "Avg S4", "Avg S5",
        "Avg CP(P)", "Avg CP(T)", "Avg AP(E)", "Avg SPC", "Avg HSM",
    ]
    for col_idx, header in enumerate(summary_headers, 1):
        ws_summary.cell(row=1, column=col_idx, value=header)
    style_header_row(ws_summary, len(summary_headers))

    job_type_data = get_job_type_manpower(records)
    row_num = 2
    for jtype in sorted(job_type_data.keys()):
        data = job_type_data[jtype]
        avg_teams = data.get("avg_team_counts", {})
        avg_roles = data.get("avg_roles", {})
        values = [
            jtype,
            data["total_jobs"],
            data["total_workers"],
            data["avg_workers"],
            avg_teams.get("S2", 0),
            avg_teams.get("S3", 0),
            avg_teams.get("S4", 0),
            avg_teams.get("S5", 0),
            avg_roles.get("CP_P", 0),
            avg_roles.get("CP_T", 0),
            avg_roles.get("AP_E", 0),
            avg_roles.get("SPC", 0),
            avg_roles.get("HSM", 0),
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_summary.cell(row=row_num, column=col_idx, value=val)
            style_data_cell(cell)
        row_num += 1

    for col_idx in range(1, len(summary_headers) + 1):
        ws_summary.column_dimensions[chr(64 + min(col_idx, 26))].width = 16

    # ── Sheet 3: Role Frequency Matrix ────────────────────────────────────
    ws_roles = workbook.create_sheet("Role Frequency")

    role_headers = ["Name", "CP(P)", "CP(T)", "AP(E)", "SPC", "HSM", "NP", "Total"]
    for col_idx, header in enumerate(role_headers, 1):
        ws_roles.cell(row=1, column=col_idx, value=header)
    style_header_row(ws_roles, len(role_headers))

    role_freq = get_role_frequency(records)
    row_num = 2
    for entry in role_freq:
        values = [
            entry["name"],
            entry.get("CP_P", 0),
            entry.get("CP_T", 0),
            entry.get("AP_E", 0),
            entry.get("SPC", 0),
            entry.get("HSM", 0),
            entry.get("NP", 0),
            entry["total"],
        ]
        for col_idx, val in enumerate(values, 1):
            cell = ws_roles.cell(row=row_num, column=col_idx, value=val)
            style_data_cell(cell)
        row_num += 1

    ws_roles.column_dimensions["A"].width = 20
    for col_idx in range(2, len(role_headers) + 1):
        ws_roles.column_dimensions[chr(64 + col_idx)].width = 10

    # ── Sheet 4: Weekly Team Distribution ─────────────────────────────────
    ws_teams = workbook.create_sheet("Weekly Team Distribution")

    team_headers = ["Week", "S2", "S3", "S4", "S5", "Total"]
    for col_idx, header in enumerate(team_headers, 1):
        ws_teams.cell(row=1, column=col_idx, value=header)
    style_header_row(ws_teams, len(team_headers))

    team_dist = get_team_distribution(records)
    row_num = 2
    for week, teams in team_dist.items():
        s2 = teams.get("S2", 0)
        s3 = teams.get("S3", 0)
        s4 = teams.get("S4", 0)
        s5 = teams.get("S5", 0)
        values = [week, s2, s3, s4, s5, s2 + s3 + s4 + s5]
        for col_idx, val in enumerate(values, 1):
            cell = ws_teams.cell(row=row_num, column=col_idx, value=val)
            style_data_cell(cell)
        row_num += 1

    for col_idx in range(1, len(team_headers) + 1):
        ws_teams.column_dimensions[chr(64 + col_idx)].width = 12

    # Save
    workbook.save(filepath)
    logger.info(f"Manpower Excel exported to {filepath}")
    return filepath


class ManpowerAnalyzer:
    """High-level analyzer wrapping all manpower analysis functions."""

    def __init__(self, records: List[Dict[str, Any]] = None):
        self.records = records or []

    def set_records(self, records: List[Dict[str, Any]]):
        self.records = records

    def summary_kpis(self) -> Dict[str, Any]:
        return get_summary_kpis(self.records)

    def daily_headcount(self) -> List[Dict[str, Any]]:
        return get_daily_headcount(self.records)

    def team_distribution(self) -> Dict[str, Dict[str, int]]:
        return get_team_distribution(self.records)

    def job_type_manpower(self) -> Dict[str, Dict[str, Any]]:
        return get_job_type_manpower(self.records)

    def role_frequency(self) -> List[Dict[str, Any]]:
        return get_role_frequency(self.records)

    def work_access_analysis(self) -> Dict[str, Dict[str, Any]]:
        return get_work_access_analysis(self.records)

    def individual_stats(self) -> List[Dict[str, Any]]:
        return get_individual_stats(self.records)

    def export_excel(self, filepath: Path) -> Path:
        return export_manpower_excel(self.records, filepath)
