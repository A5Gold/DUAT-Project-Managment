# MTR PS-OHLR DUAT - S-Curve Generation
"""
S-Curve calculation and plotting for project progress visualization.
"""

import pandas as pd
import numpy as np
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import logging
from datetime import datetime
from typing import Dict, List, Any, Optional, Tuple
from pathlib import Path
import io
import base64

logger = logging.getLogger(__name__)


def calculate_scurve_data(
    df: pd.DataFrame,
    project_code: str,
    target_qty: float,
    start_year: int,
    start_week: int,
    end_year: int,
    end_week: int
) -> Tuple[List[str], List[float], List[float], float]:
    """
    Calculate S-Curve data for a specific project.
    
    Args:
        df: DataFrame with project data
        project_code: Project code (e.g., "C2264")
        target_qty: Target quantity
        start_year: Start year
        start_week: Start week
        end_year: End year
        end_week: End week
        
    Returns:
        Tuple of (week_labels, cumulative_target, cumulative_actual, current_progress_pct)
    """
    # Filter data for this project
    project_df = df[df['Project'].str.upper() == project_code.upper()].copy()
    
    if project_df.empty:
        return [], [], [], 0.0
    
    # Calculate total weeks (handle years with 53 ISO weeks)
    def _weeks_in_year(year: int) -> int:
        """Return 52 or 53 depending on ISO week count for the year."""
        from datetime import date
        dec28 = date(year, 12, 28)
        return dec28.isocalendar()[1]

    total_weeks = 0
    y, w = start_year, start_week
    while (y, w) != (end_year, end_week):
        total_weeks += 1
        w += 1
        if w > _weeks_in_year(y):
            w = 1
            y += 1
    if total_weeks <= 0:
        return [], [], [], 0.0

    # Generate week labels
    week_labels = []
    current_year = start_year
    current_week = start_week

    for _ in range(total_weeks + 1):
        week_labels.append(f"{current_year}-W{current_week:02d}")
        current_week += 1
        if current_week > _weeks_in_year(current_year):
            current_week = 1
            current_year += 1
    
    # Calculate cumulative target (linear distribution)
    weekly_target = target_qty / total_weeks
    cumulative_target = [weekly_target * (i + 1) for i in range(len(week_labels))]
    
    # Calculate cumulative actual from data
    project_df['YearWeek'] = project_df['Year'].astype(str) + '-W' + project_df['Week'].astype(str).str.zfill(2)
    weekly_actual = project_df.groupby('YearWeek')['Qty Delivered'].sum()
    
    cumulative_actual = []
    running_total = 0
    for week in week_labels:
        if week in weekly_actual.index:
            running_total += weekly_actual[week]
        cumulative_actual.append(running_total)
    
    # Calculate progress percentage
    current_progress = (cumulative_actual[-1] / target_qty * 100) if target_qty > 0 and cumulative_actual else 0
    
    return week_labels, cumulative_target, cumulative_actual, round(current_progress, 1)


def plot_scurve(
    week_labels: List[str],
    cumulative_target: List[float],
    cumulative_actual: List[float],
    project_code: str,
    target_qty: float,
    output_path: Path = None,
    figsize: Tuple[int, int] = (12, 6)
) -> Optional[str]:
    """
    Plot S-Curve chart.
    
    Args:
        week_labels: List of week labels
        cumulative_target: Cumulative target values
        cumulative_actual: Cumulative actual values
        project_code: Project code for title
        target_qty: Target quantity for title
        output_path: Optional path to save the figure
        figsize: Figure size
        
    Returns:
        Base64 encoded image string if no output_path, else None
    """
    fig, ax = plt.subplots(figsize=figsize)
    
    # Plot target line
    ax.plot(
        range(len(week_labels)), 
        cumulative_target, 
        'b--', 
        linewidth=2, 
        label='Target (Linear)',
        alpha=0.7
    )
    
    # Plot actual line
    ax.plot(
        range(len(week_labels)), 
        cumulative_actual, 
        'g-', 
        linewidth=2.5, 
        label='Actual Progress',
        marker='o',
        markersize=4
    )
    
    # Fill area between
    ax.fill_between(
        range(len(week_labels)),
        cumulative_actual,
        cumulative_target,
        alpha=0.2,
        color='green' if cumulative_actual[-1] >= cumulative_target[-1] else 'red'
    )
    
    # Styling
    ax.set_xlabel('Week', fontsize=10)
    ax.set_ylabel('Cumulative Quantity', fontsize=10)
    ax.set_title(f'S-Curve: {project_code} (Target: {target_qty:,.0f})', fontsize=12, fontweight='bold')
    ax.legend(loc='upper left')
    ax.grid(True, alpha=0.3)
    
    # X-axis labels (show every Nth label)
    n_labels = len(week_labels)
    step = max(1, n_labels // 12)
    ax.set_xticks(range(0, n_labels, step))
    ax.set_xticklabels([week_labels[i] for i in range(0, n_labels, step)], rotation=45, ha='right')
    
    plt.tight_layout()
    
    if output_path:
        fig.savefig(output_path, dpi=150, bbox_inches='tight')
        plt.close(fig)
        return None
    else:
        # Return base64 encoded image
        buf = io.BytesIO()
        fig.savefig(buf, format='png', dpi=150, bbox_inches='tight')
        buf.seek(0)
        img_base64 = base64.b64encode(buf.read()).decode('utf-8')
        plt.close(fig)
        return img_base64


def generate_scurve_excel(
    df: pd.DataFrame,
    project_code: str,
    target_qty: float,
    start_year: int,
    start_week: int,
    end_year: int,
    end_week: int,
    output_path: Path
) -> Tuple[bool, float]:
    """
    Generate S-Curve Excel report with embedded chart.
    
    Args:
        df: DataFrame with project data
        project_code: Project code
        target_qty: Target quantity
        start_year, start_week: Start period
        end_year, end_week: End period
        output_path: Output file path
        
    Returns:
        Tuple of (success, progress_percentage)
    """
    from openpyxl import Workbook
    from openpyxl.drawing.image import Image as XLImage
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    
    try:
        # Calculate S-Curve data
        week_labels, cumulative_target, cumulative_actual, progress = calculate_scurve_data(
            df, project_code, target_qty, start_year, start_week, end_year, end_week
        )
        
        if not week_labels:
            return False, 0.0
        
        # Create workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "S-Curve Data"
        
        # Headers
        headers = ["Week", "Target (Cumulative)", "Actual (Cumulative)", "Variance"]
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
            cell.font = Font(bold=True, color="FFFFFF")
            cell.alignment = Alignment(horizontal="center")
        
        # Data rows
        for row, (week, target, actual) in enumerate(zip(week_labels, cumulative_target, cumulative_actual), 2):
            ws.cell(row=row, column=1, value=week)
            ws.cell(row=row, column=2, value=round(target, 2))
            ws.cell(row=row, column=3, value=round(actual, 2))
            ws.cell(row=row, column=4, value=round(actual - target, 2))
        
        # Auto-adjust column widths
        for col in range(1, 5):
            ws.column_dimensions[chr(64 + col)].width = 18
        
        # Generate chart image
        chart_path = output_path.parent / f"temp_scurve_{project_code}.png"
        plot_scurve(week_labels, cumulative_target, cumulative_actual, project_code, target_qty, chart_path)
        
        # Add chart image to a new sheet
        if chart_path.exists():
            ws_chart = wb.create_sheet("S-Curve Chart")
            img = XLImage(str(chart_path))
            ws_chart.add_image(img, "A1")

        wb.save(output_path)

        # Clean up temp file after save
        if chart_path.exists():
            chart_path.unlink()
        return True, progress
        
    except Exception as e:
        logger.error(f"Error generating S-Curve Excel: {e}")
        return False, 0.0


class SCurveGenerator:
    """S-Curve generator class."""
    
    def __init__(self, df: pd.DataFrame = None):
        self.df = df
    
    def set_data(self, df: pd.DataFrame):
        """Set the data DataFrame."""
        self.df = df
    
    def generate(
        self,
        project_code: str,
        target_qty: float,
        start_year: int,
        start_week: int,
        end_year: int,
        end_week: int,
        output_path: Path
    ) -> Tuple[bool, float]:
        """
        Generate S-Curve report.
        
        Returns:
            Tuple of (success, progress_percentage)
        """
        if self.df is None or self.df.empty:
            return False, 0.0
        
        return generate_scurve_excel(
            self.df, project_code, target_qty,
            start_year, start_week, end_year, end_week,
            output_path
        )
