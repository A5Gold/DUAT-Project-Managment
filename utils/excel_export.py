# MTR DUAT - Excel Export Utilities
"""
Functions for exporting pandas DataFrames to formatted Excel workbooks.

Used by:
- backend/routers/export.py  (dashboard + lag analysis exports)
- analysis/dashboard.py      (DashboardAnalyzer.export)
"""

import logging
from pathlib import Path
from typing import Optional, Union

import pandas as pd
from openpyxl.utils import get_column_letter

logger = logging.getLogger(__name__)

PathLike = Union[str, Path]


def _auto_adjust_column_widths(writer: pd.ExcelWriter, sheet_name: str) -> None:
    """Widen every column to fit its longest value (header included)."""
    ws = writer.sheets[sheet_name]
    for col_idx, col_cells in enumerate(ws.iter_cols(min_row=1), start=1):
        max_length = 0
        for cell in col_cells:
            try:
                cell_len = len(str(cell.value)) if cell.value is not None else 0
                if cell_len > max_length:
                    max_length = cell_len
            except (TypeError, AttributeError):
                pass
        # Add a small padding so text doesn't touch the cell border
        adjusted = max_length + 2
        ws.column_dimensions[get_column_letter(col_idx)].width = adjusted


def export_dataframe_to_excel(df: pd.DataFrame, path: PathLike) -> bool:
    """
    Export a single DataFrame to an Excel file.

    Args:
        df: DataFrame to export.
        path: Destination file path.

    Returns:
        True on success, False on any error.
    """
    try:
        path = Path(path)
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Sheet1", index=False)
        return True
    except Exception as exc:
        logger.error("export_dataframe_to_excel failed: %s", exc)
        return False


def create_dashboard_excel(
    raw_df: pd.DataFrame,
    summary_df: pd.DataFrame,
    path: PathLike,
    nth_trend: Optional[pd.DataFrame] = None,
) -> bool:
    """
    Create a multi-sheet dashboard Excel workbook.

    Sheets created:
        1. "Raw Data"             - raw_df
        2. "Summary by Project"   - summary_df
        3. "NTH Trend by Week"    - nth_trend (only when provided and non-empty)

    Args:
        raw_df: Raw data DataFrame.
        summary_df: Summary statistics DataFrame.
        path: Destination file path.
        nth_trend: Optional NTH pivot table (index = YearWeek).

    Returns:
        True on success, False on any error.
    """
    try:
        path = Path(path)
        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            raw_df.to_excel(writer, sheet_name="Raw Data", index=False)
            _auto_adjust_column_widths(writer, "Raw Data")

            summary_df.to_excel(writer, sheet_name="Summary by Project", index=False)
            _auto_adjust_column_widths(writer, "Summary by Project")

            if nth_trend is not None and not nth_trend.empty:
                nth_trend.to_excel(writer, sheet_name="NTH Trend by Week", index=True)
                _auto_adjust_column_widths(writer, "NTH Trend by Week")

        return True
    except Exception as exc:
        logger.error("create_dashboard_excel failed: %s", exc)
        return False


def export_lag_analysis_report(results_df: pd.DataFrame, path: PathLike) -> bool:
    """
    Export lag analysis results to an Excel file.

    The "Status Color" column is dropped before writing.  Column widths are
    auto-adjusted so the report is immediately readable.

    Args:
        results_df: Lag analysis results DataFrame.
        path: Destination file path.

    Returns:
        True on success, False on any error.
    """
    try:
        path = Path(path)
        # Immutable: work on a copy, never touch the caller's DataFrame
        df = results_df.copy()
        if "Status Color" in df.columns:
            df = df.drop(columns=["Status Color"])

        with pd.ExcelWriter(path, engine="openpyxl") as writer:
            df.to_excel(writer, sheet_name="Lag Analysis", index=False)
            _auto_adjust_column_widths(writer, "Lag Analysis")

        return True
    except Exception as exc:
        logger.error("export_lag_analysis_report failed: %s", exc)
        return False
